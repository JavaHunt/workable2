from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time, json, pyperclip
import openpyxl


# Load the existing workbook
try:
    workbook = openpyxl.load_workbook("data.xlsx")
    print('data.xlsx is already exists')
except FileNotFoundError:
    workbook =  openpyxl.Workbook()
    # Select the active sheet
    sheet = workbook.active
    sheet.append(list(('name', 'email', 'phone', 'created at', 'resume pdf', 'resume url')))

sheet = workbook.active

# users_url = input("Enter the link of users : ")

first_url = 'https://hyperverge.workable.com/backend/api/candidates/first?job_id=3200225'

start = time.time()

job_id = first_url[first_url.find('=') + 1:]

# Open Microsoft Edge browser
driver = webdriver.Edge()

# To use chrome try this
# driver = webdriver.Chrome()

# Navigate to the login page
driver.get('https://id.workable.com/oidc/i/PmvTD7Xu0R7C5fFRQlIBZ')  # Replace with your login page URL

# Find the username and password input fields and enter the credentials
username_field = driver.find_element(By.ID, 'email')  # Replace with the ID or other locator of the username field
password_field = driver.find_element(By.ID, 'password')  # Replace with the ID or other locator of the password field

username_field.send_keys('pardhu9100@gmail.com')  # Replace with your username
password_field.send_keys('P@rdhu08092002')  # Replace with your password

# Submit the login form
password_field.send_keys(Keys.ENTER)

# Wait for the login process to complete
time.sleep(17)  # Adjust the waiting time if needed

# Open a URL after the login process
driver.get(first_url)

time.sleep(4)

# Find the body element or the HTML element
body = driver.find_element(By.TAG_NAME, 'body')

# Simulate pressing Ctrl + A on the body or HTML element
body.send_keys(Keys.CONTROL, 'a')
time.sleep(0.5)
body.send_keys(Keys.CONTROL, 'c')


# Get the copied content from the clipboard
copied_content = pyperclip.paste()

first_user_filename = 'first_user.json'
user_filename = 'users_list.json'
first_user_id = ''


# Store the webpage content in a JSON file
with open(first_user_filename, 'w', encoding="utf-8") as file:
    file.write(copied_content)


# now we got the list
# Read the JSON file

with open(first_user_filename, 'r', encoding='utf-8') as file:
    json_data = json.load(file)


no_of_users = 0  # Total candidates
count = 0
# Check if data is not in the URl
if(len(json_data) == 0):
    print(f"No data in the provided URL: {first_url}  Check the URL you provided...")

else:
    # This will add new row in the excel sheet
    sheet.append(list((json_data['name'], json_data['email'], json_data['phone'], json_data['created_at'], 
                      json_data['resume_pdf_url'], json_data['resume_url'])))
    workbook.save("data.xlsx")
    first_user_id = json_data['id']
    # print(" user ", str(len(data['name'])), ' / ', str(no_of_users), " completed... with id = ", str(first_user_id))

def get_users_data(user_id):
    global no_of_users, count, sheet
    
    users_url = f'https://hyperverge.workable.com/backend/api/candidates/list?job_id={job_id}&older_than={str(user_id)}'
    driver.get(users_url)

    # Use this print statement for debugging to know which list is currently running
    print('users url in loop :', users_url, '  ', sheet.max_row-1)
    time.sleep(1.8) # decrease sleep time if your internet speed is high

    # Find the body element or the HTML element
    body = driver.find_element(By.TAG_NAME, 'body')

    # Simulate pressing Ctrl + A on the body or HTML element
    body.send_keys(Keys.CONTROL, 'a')
    time.sleep(0.1)
    body.send_keys(Keys.CONTROL, 'c')


    # Get the copied content from the clipboard
    copied_content = pyperclip.paste()

    # Store the webpage content in a JSON file
    with open(user_filename, 'w', encoding="utf-8") as file:
        file.write(copied_content)
    

    with open(user_filename, 'r', encoding='utf-8') as file:
        json_data = json.load(file)
    
    # Get no of users
    count += 1
    if(count == 1):
        # checking of no of users once is enough
        no_of_users = json_data['total']
    
    if(len(json_data) != 0 and sheet.max_row <= no_of_users):
        try:
            for i in range(len(json_data['data'])):
                data = []  # temp list to save user details
                user_url_by_id = "https://hyperverge.workable.com/backend/api/candidates/"
                data.append(json_data['data'][i]['name'])
                data.append(json_data['data'][i]['email'])
                data.append(json_data['data'][i]['phone'])
                data.append(json_data['data'][i]['created_at'])
                id = json_data['data'][i]['id']
                user_url_by_id += str(id)
                driver.get(user_url_by_id)  # This will navigate to the user details
                time.sleep(1.5)

                # Find the body element or the HTML element
                body = driver.find_element(By.TAG_NAME, 'body')

                # Simulate pressing Ctrl + A on the body or HTML element
                body.send_keys(Keys.CONTROL, 'a')
                time.sleep(0.1)
                body.send_keys(Keys.CONTROL, 'c')

                # Get the copied content from the clipboard
                copied_content = pyperclip.paste()

                # Store the webpage content in a temporary JSON file
                with open('temp_user.json', 'w', encoding="utf-8") as file:
                    file.write(copied_content)
                
                with open('temp_user.json', 'r', encoding='utf-8') as file:
                    user_data = json.load(file)

                try:
                    data.append(user_data['resume_pdf_url'])
                except KeyError:
                    data.append(None)
                try:
                    data.append(user_data['resume_url'])
                except:
                    data.append(None)
                sheet.append(data)
                workbook.save('data.xlsx')
                current_time = time.time()
                print(" user ", str(sheet.max_row-1), ' / ', str(no_of_users), " completed with id = ", str(id), '  ',
                    (current_time-start)/60, ' minutes')
        except KeyError as e:
            print(e)
            print('user id : ', user_id)
            print('users url :', users_url)
        # Recursive until the last User
        time.sleep(0.3)
        get_users_data(json_data['data'][-1]['id'])
    else:
        try:
            workbook.save('data.xlsx')
            print(" All Users details Added to 'data.xlsx' Excel file")
        except PermissionError:
            print("Close the excel of 'data.xlsx' if you opened.... If not, give read and write permissions to it in it's properties")

            
        
get_users_data(first_user_id)
time.sleep(2)

end = time.time()
print("\nRunning time : ", (end-start)/60, ' minutes')